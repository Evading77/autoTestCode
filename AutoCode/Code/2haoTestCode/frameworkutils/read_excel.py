import re

import openpyxl

# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('F:\employee.xlsx')
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单
# print('{}行 {}列'.format(ws.max_row, ws.max_column))

# print(ws.cell(row=1, column=2)) # 获取第一行第二列的单元格
# print(ws.cell(row=1, column=2).value)
# for i in range(1, 8, 2): #  获取1,3,4,7 行第二列的值
#     print(i, ws.cell(row=i, column=2).value)
# col_range = ws['A:B']
# row_range = ws[1:3]

# for col in col_range:  # 打印BC两列单元格中的值内容
#     for cell in col:
#         print(cell.value)
#
# for row in row_range:  # 打印 2-5行中所有单元格中的值
#     for cell in row:
#         print(cell.value)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=2):
    for cell in row:
        if "assert" in cell.value:
            assertValue=re.split(":",cell.value)[1]
            print(assertValue)


        print("haha")