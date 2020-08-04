import os
import re
import time

import openpyxl


class ThirdPageCheck(object):
    #��ת�����˵��������Է���
    def third_pages_check(self):
        # ��excel�ļ�,��ȡ����������
        filePath=os.path.dirname(os.path.abspath('.')) + '/config/employee.xlsx'
        wb = openpyxl.load_workbook(filePath)
        # �ӱ��л�ȡ��Ԫ�������
        ws = wb.active  # ��ǰ��Ծ�ı�
        for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=2):  # ��ӡ1-2�У�1-2���е�����
            for cell in row:
                xpath_value = str(cell.value)
                if "assert:" in cell.value:
                    assertXpath = 'xpath=>' + re.split(":", cell.value)[1]
                    self.assert_ele_is_exist(assertXpath)
                else:
                    self.driver.find_element_by_xpath(xpath_value).click()
                    time.sleep(1)